from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.utils.excel_dates import convert_excel_date
from src.utils.text import error_type, is_error_value


@dataclass
class CellIssue:
    arquivo: str
    aba: str
    celula: str
    valor_erro: str
    formula: str | None
    tipo_erro: str
    possivel_impacto: str
    recomendacao: str


def load_workbook_pair(file_path: str | Path):
    path = Path(file_path)
    formula_wb = load_workbook(path, data_only=False, read_only=False)
    value_wb = load_workbook(path, data_only=True, read_only=False)
    return formula_wb, value_wb


def used_range(min_row: int | None, min_col: int | None, max_row: int | None, max_col: int | None) -> str:
    if None in {min_row, min_col, max_row, max_col}:
        return "A1:A1"
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def inspect_workbook(file_path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    formula_wb, value_wb = load_workbook_pair(file_path)
    file_name = Path(file_path).name
    catalog_rows: list[dict[str, Any]] = []
    error_rows: list[dict[str, Any]] = []

    for sheet_name in formula_wb.sheetnames:
        ws_formula = formula_wb[sheet_name]
        ws_values = value_wb[sheet_name]
        non_empty = 0
        formula_count = 0
        error_count = 0
        min_row = min_col = max_row = max_col = None

        for row in range(1, ws_formula.max_row + 1):
            for col in range(1, ws_formula.max_column + 1):
                cell_formula = ws_formula.cell(row, col)
                cell_value = ws_values.cell(row, col)
                raw_formula = cell_formula.value
                raw_value = cell_value.value
                if raw_formula is None and raw_value is None:
                    continue
                non_empty += 1
                min_row = row if min_row is None else min(min_row, row)
                min_col = col if min_col is None else min(min_col, col)
                max_row = row if max_row is None else max(max_row, row)
                max_col = col if max_col is None else max(max_col, col)
                if isinstance(raw_formula, str) and raw_formula.startswith("="):
                    formula_count += 1
                possible_error = raw_value if raw_value is not None else raw_formula
                if is_error_value(possible_error):
                    error_count += 1
                    err = error_type(possible_error) or "ERRO"
                    impact = "Pode comprometer KPIs e gráficos dependentes desta célula."
                    if err == "#REF!":
                        impact = "Referência quebrada: indicador ou total pode estar incorreto."
                    elif err == "#DIV/0!":
                        impact = "Divisão por zero: taxa/percentual fica inválido e precisa de contexto."
                    error_rows.append(
                        {
                            "arquivo": file_name,
                            "aba": sheet_name,
                            "celula": cell_formula.coordinate,
                            "valor_erro": str(possible_error),
                            "formula": raw_formula if isinstance(raw_formula, str) and raw_formula.startswith("=") else None,
                            "tipo_erro": err,
                            "possivel_impacto": impact,
                            "recomendacao": "Revisar a fórmula e a origem referenciada antes de usar no dashboard.",
                        }
                    )

        catalog_rows.append(
            {
                "arquivo": file_name,
                "aba": sheet_name,
                "dimensao_usada": used_range(min_row, min_col, max_row, max_col),
                "linhas_max": ws_formula.max_row,
                "colunas_max": ws_formula.max_column,
                "celulas_preenchidas": non_empty,
                "quantidade_formulas": formula_count,
                "quantidade_erros": error_count,
            }
        )

    return catalog_rows, error_rows


def sheet_records(file_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    formula_wb, value_wb = load_workbook_pair(file_path)
    ws_formula = formula_wb[sheet_name]
    ws_values = value_wb[sheet_name]
    records: list[dict[str, Any]] = []
    for row in range(1, ws_formula.max_row + 1):
        for col in range(1, ws_formula.max_column + 1):
            cell_formula = ws_formula.cell(row, col)
            cell_value = ws_values.cell(row, col)
            formula = cell_formula.value if isinstance(cell_formula.value, str) and cell_formula.value.startswith("=") else None
            value = cell_value.value if cell_value.value is not None else cell_formula.value
            records.append(
                {
                    "row": row,
                    "column": col,
                    "coordinate": cell_formula.coordinate,
                    "formula": formula,
                    "value": convert_excel_date(value),
                }
            )
    return records


def worksheet_to_table(file_path: str | Path, sheet_name: str, header_row: int = 1) -> list[dict[str, Any]]:
    formula_wb, value_wb = load_workbook_pair(file_path)
    ws_formula = formula_wb[sheet_name]
    ws_values = value_wb[sheet_name]
    headers = []
    for col in range(1, ws_formula.max_column + 1):
        header = ws_formula.cell(header_row, col).value
        headers.append(header if header is not None else f"col_{col}")
    rows: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws_formula.max_row + 1):
        record = {}
        has_value = False
        for col, header in enumerate(headers, start=1):
            value = ws_values.cell(row, col).value
            if value is None:
                value = ws_formula.cell(row, col).value
            value = convert_excel_date(value)
            record[str(header)] = value
            has_value = has_value or value not in (None, "")
        if has_value:
            rows.append(record)
    return rows

